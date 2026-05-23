from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from rag.ingest.asset_anchors import asset_anchor
from rag.ingest.parsers.excel_parser_repo import ExcelParserRepo
from rag.schema.core import SourceType


def test_excel_parser_schema_samples_large_sheet_without_full_markdown(tmp_path) -> None:
    source_path = tmp_path / "ledger.xlsx"
    rows = [{"Name": f"row{index}", "Department": f"dept{index % 3}", "Amount": index * 100} for index in range(800)]
    pd.DataFrame(rows).to_excel(source_path, sheet_name="Ledger", index=False)

    parsed = ExcelParserRepo().parse(
        source_path,
        location=str(source_path),
        source_type=SourceType.XLSX,
        title="Ledger",
        owner="tester",
    )

    assert len(parsed.sections) == 1
    assert len(parsed.elements) == 1
    element = parsed.elements[0]
    assert element.kind == "table"
    assert element.metadata["sheet_name"] == "Ledger"
    assert element.metadata["row_count"] == 800
    assert element.metadata["column_count"] == 3
    assert element.metadata["table_policy"] == "compute_only"
    assert element.metadata["sample_rows"][0] == {
        "Name": "row0",
        "Department": "dept0",
        "Amount": "0",
    }
    assert element.metadata["schema"][2]["name"] == "Amount"
    assert element.metadata["schema"][2]["type"].startswith("number")
    assert "Table columns: Name | Department | Amount" in element.text
    assert "row0" in element.text
    assert "row799" not in element.text
    assert asset_anchor(element.element_id) in parsed.sections[0].text


def test_excel_parser_preserves_full_shape_when_profile_reads_are_bounded(tmp_path) -> None:
    source_path = tmp_path / "large-ledger.xlsx"
    row_count = 5_010
    rows = [
        {"Name": f"row{index}", "Department": f"dept{index % 3}", "Amount": index * 100}
        for index in range(row_count)
    ]
    pd.DataFrame(rows).to_excel(source_path, sheet_name="Ledger", index=False)

    parsed = ExcelParserRepo().parse(
        source_path,
        location=str(source_path),
        source_type=SourceType.XLSX,
        title="Large Ledger",
        owner="tester",
    )

    element = parsed.elements[0]
    assert element.metadata["row_count"] == row_count
    assert element.metadata["column_count"] == 3
    assert element.metadata["profile_rows_read"] <= 5_000
    assert element.metadata["sample_rows"][0] == {
        "Name": "row0",
        "Department": "dept0",
        "Amount": "0",
    }
    assert "Table shape: rows=5010, columns=3" in element.text


def test_excel_parser_ignores_trailing_styled_blank_rows_when_sheet_is_fully_previewed(tmp_path) -> None:
    source_path = tmp_path / "styled-blanks.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ledger"
    worksheet.append(["Name", "Amount"])
    worksheet.append(["Travel", 500])
    worksheet["A20"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    workbook.save(source_path)

    parsed = ExcelParserRepo().parse(
        source_path,
        location=str(source_path),
        source_type=SourceType.XLSX,
        title="Styled Blanks",
        owner="tester",
    )

    element = parsed.elements[0]
    assert element.metadata["row_count"] == 1
    assert element.metadata["profile_rows_read"] == 1
    assert element.metadata["row_count_source"] == "preview"


def test_excel_parser_evaluates_common_formula_cells_without_cached_values(tmp_path) -> None:
    source_path = tmp_path / "formula-report.xlsx"
    workbook = Workbook()
    report = workbook.active
    report.title = "Report"
    report.append(["Key", "Group", "Amount", "GroupTotal", "LookupAmount"])
    report.append(["north-a", "North", 10, '=SUMIF($B$2:$B$3,$B2,$C$2:$C$3)', '=IFERROR(VLOOKUP($A2,Data!$A:$B,2,0),0)'])
    report.append(["north-b", "North", 20, '=SUMIF($B$2:$B$3,$B3,$C$2:$C$3)', '=IFERROR(VLOOKUP($A3,Data!$A:$B,2,0),0)'])
    report.append(["total", "All", '=SUM(C2:C3)', '=SUM(D2:D3)', ""])
    data = workbook.create_sheet("Data")
    data.append(["Key", "LookupAmount"])
    data.append(["north-a", 100])
    data.append(["north-b", 200])
    workbook.save(source_path)

    parsed = ExcelParserRepo().parse(
        source_path,
        location=str(source_path),
        source_type=SourceType.XLSX,
        title="Formula Report",
        owner="tester",
    )

    report_element = parsed.elements[0]
    assert report_element.metadata["sample_rows"][0] == {
        "Key": "north-a",
        "Group": "North",
        "Amount": "10",
        "GroupTotal": "30",
        "LookupAmount": "100",
    }
    assert report_element.metadata["sample_rows"][2]["Amount"] == "30"
    assert report_element.metadata["sample_rows"][2]["GroupTotal"] == "60"
