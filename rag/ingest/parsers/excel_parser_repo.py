from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from rag.ingest.asset_anchors import asset_anchor
from rag.ingest.excel_formula_evaluator import ExcelFormulaEvaluator
from rag.ingest.header_detector import MAX_SCAN_ROWS, HeaderKind, detect_header
from rag.ingest.parsers.util import default_title_from_location, normalize_whitespace, slugify
from rag.ingest.table_sampler import profile_table_data
from rag.schema.core import ParsedDocument, ParsedElement, ParsedSection, SourceType

_MIN_DETECT_CONFIDENCE = 0.5

# 大文件保护：只读取用于表头识别和摘要画像的有限预览
_MAX_SHEET_ROWS = 5_000
_MAX_SHEET_COLUMNS = 50


class ExcelParserRepo:
    """
    专门针对 Excel (XLSX/XLS) 格式的解析特种兵。
    核心使命：拒绝视觉渲染，直接读取底层数据，保住极其脆弱的二维表格关系。
    """

    def parse(
        self,
        file_path: Path,
        *,
        location: str,
        source_type: SourceType,
        title: str | None = None,
        owner: str = "user",
    ) -> ParsedDocument:
        doc_title = title or default_title_from_location(location)

        formula_workbook = None
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            formula_workbook = load_workbook(file_path, read_only=True, data_only=False)
            formula_evaluator = ExcelFormulaEvaluator(
                value_workbook=workbook,
                formula_workbook=formula_workbook,
            )
            sheet_names = workbook.sheetnames
        except Exception as exc:
            raise ValueError(f"Excel parser failed to read {location}: {exc}") from exc

        try:
            sections: list[ParsedSection] = []
            all_elements: list[ParsedElement] = []
            visible_text_parts: list[str] = []
            visible_cursor = 0
            visible_section_separator = "\n\n"

            valid_section_count = 0
            total_sheets = len(sheet_names)

            for sheet_idx, sheet_name in enumerate(sheet_names):
                worksheet = formula_workbook[sheet_name]
                (
                    raw_df,
                    total_sheet_rows,
                    total_column_count,
                    profile_columns_read,
                    row_count_source,
                ) = self._read_sheet_preview(worksheet, evaluator=formula_evaluator)

                if raw_df.empty:
                    continue

                clean_sheet_name = normalize_whitespace(sheet_name)
                element_id = f"{slugify(doc_title)}-sheet-{sheet_idx}-table"

                header_result = detect_header(raw_df)

                columns: list[str]
                rows: list[list[str]]
                if header_result.confidence >= _MIN_DETECT_CONFIDENCE and header_result.header_kind != HeaderKind.NONE:
                    columns = [
                        normalize_whitespace(col) or f"column_{i + 1}"
                        for i, col in enumerate(header_result.normalized_columns)
                    ]
                    data_start = header_result.data_start_row
                else:
                    columns = [f"column_{column_index + 1}" for column_index in range(raw_df.shape[1])]
                    data_start = 0

                rows = [
                    [self._cell_text(value) for value in row]
                    for row in raw_df.iloc[data_start : data_start + _MAX_SHEET_ROWS].itertuples(
                        index=False,
                        name=None,
                    )
                ]
                rows = [row for row in rows if any(cell.strip() for cell in row)]

                # 3. 大表保护：画像只用有界预览，但 metadata 保留 sheet 的完整规模。
                sampled_rows = self._sample_rows(rows, columns)
                sampled_columns = columns[:_MAX_SHEET_COLUMNS]
                sampled_rows = [row[: len(sampled_columns)] for row in sampled_rows]
                total_data_rows = max(total_sheet_rows - data_start, len(rows))

                table_profile = profile_table_data(
                    columns=sampled_columns,
                    rows=sampled_rows,
                    total_row_count=total_data_rows,
                    total_column_count=total_column_count,
                )
                asset_text = table_profile.summary_sample

                # 4. 封装 Element
                all_elements.append(
                    ParsedElement(
                        element_id=element_id,
                        kind="table",
                        text=asset_text,
                        toc_path=(doc_title, clean_sheet_name),
                        page_no=sheet_idx + 1,
                        metadata={
                            "source_type": SourceType.XLSX.value,
                            "sheet_name": clean_sheet_name,
                            "row_count": table_profile.row_count,
                            "column_count": table_profile.column_count,
                            "estimated_tokens": table_profile.estimated_tokens,
                            "table_policy": table_profile.table_policy,
                            "asset_summary_sample": table_profile.summary_sample,
                            "asset_text_preview": table_profile.preview_text,
                            "sample_rows": table_profile.sample_rows,
                            "schema": table_profile.schema,
                            "header_kind": header_result.header_kind.name,
                            "header_confidence": header_result.confidence,
                            "profile_rows_read": len(sampled_rows),
                            "profile_columns_read": profile_columns_read,
                            "row_count_source": row_count_source,
                            "column_count_source": "preview" if row_count_source == "preview" else "header_scan",
                        },
                    )
                )

                section_text = f"## Sheet: {clean_sheet_name}\n\n{asset_anchor(element_id)}"

                if visible_text_parts:
                    visible_text_parts.append(visible_section_separator)
                    visible_cursor += len(visible_section_separator)

                char_range_start = visible_cursor
                visible_text_parts.append(section_text)
                visible_cursor += len(section_text)
                char_range_end = visible_cursor

                # 5. 封装 Section
                sections.append(
                    ParsedSection(
                        toc_path=(doc_title, clean_sheet_name),
                        heading_level=2,
                        page_range=(sheet_idx + 1, sheet_idx + 1),
                        order_index=valid_section_count,
                        text=section_text,
                        char_range_start=char_range_start,
                        char_range_end=char_range_end,
                        anchor_hint=slugify(f"sheet-{clean_sheet_name}"),
                        metadata={"sheet_name": clean_sheet_name},
                    )
                )

                valid_section_count += 1

            visible_text = "".join(visible_text_parts)

            for idx, section in enumerate(sections):
                start = section.char_range_start
                end = section.char_range_end
                if start is None or end is None:
                    raise ValueError(f"xlsx section[{idx}] missing char range")
                if start < 0 or end <= start or end > len(visible_text):
                    raise ValueError(
                        f"xlsx section[{idx}] invalid char range: "
                        f"start={start}, end={end}, visible_len={len(visible_text)}"
                    )
                if visible_text[start:end] != section.text:
                    raise ValueError(
                        f"xlsx section[{idx}] text/span mismatch: "
                        f"expected={section.text!r}, actual={visible_text[start:end]!r}"
                    )

            return ParsedDocument(
                title=doc_title,
                source_type=SourceType.XLSX,
                authors=[owner],
                language=None,
                sections=sections,
                visible_text=visible_text,
                visual_semantics=None,
                elements=all_elements,
                page_count=total_sheets,  # 记录物理总数，无论是否为空
                metadata={
                    "location": location,
                    "source_type": SourceType.XLSX.value,
                    "valid_sections": str(valid_section_count),
                },
            )
        finally:
            workbook.close()
            if formula_workbook is not None:
                formula_workbook.close()

    def _read_sheet_preview(
        self,
        worksheet: Any,
        *,
        evaluator: ExcelFormulaEvaluator | None = None,
    ) -> tuple[pd.DataFrame, int, int, int, str]:
        max_row = int(worksheet.max_row or 0)
        max_column = int(worksheet.max_column or 0)
        if max_row <= 0 or max_column <= 0:
            return pd.DataFrame(), 0, 0, 0, "preview"

        scan_evaluator = (
            evaluator
            if self._worksheet_has_formulas(
                worksheet,
                max_row=min(max_row, MAX_SCAN_ROWS),
                max_column=max_column,
            )
            else None
        )
        scan_rows = self._worksheet_rows(
            worksheet,
            max_row=min(max_row, MAX_SCAN_ROWS),
            max_column=max_column,
            evaluator=scan_evaluator,
        )
        active_column_indices = self._active_column_indices(scan_rows)
        if not active_column_indices:
            return pd.DataFrame(), 0, 0, 0, "preview"

        total_column_count = len(active_column_indices)
        preview_column_indices = active_column_indices[:_MAX_SHEET_COLUMNS]
        preview_row_count = min(max_row, _MAX_SHEET_ROWS + MAX_SCAN_ROWS)
        row_count_source = "preview" if preview_row_count >= max_row else "worksheet_dimension"
        preview_max_column = preview_column_indices[-1] + 1
        preview_evaluator = (
            evaluator
            if self._worksheet_has_formulas(
                worksheet,
                max_row=preview_row_count,
                max_column=preview_max_column,
                column_indices=preview_column_indices,
            )
            else None
        )
        preview_rows = self._worksheet_rows(
            worksheet,
            max_row=preview_row_count,
            max_column=preview_max_column,
            column_indices=preview_column_indices,
            evaluator=preview_evaluator,
        )
        raw_df = pd.DataFrame(preview_rows)
        if raw_df.empty:
            return raw_df, 0, 0, 0, "preview"
        raw_df = raw_df.dropna(how="all").dropna(axis=1, how="all").fillna("")
        raw_df.columns = range(raw_df.shape[1])
        if row_count_source == "preview":
            total_column_count = raw_df.shape[1]
            total_sheet_rows = raw_df.shape[0]
        else:
            total_sheet_rows = max_row
        return raw_df, total_sheet_rows, total_column_count, raw_df.shape[1], row_count_source

    @classmethod
    def _worksheet_rows(
        cls,
        worksheet: Any,
        *,
        max_row: int,
        max_column: int,
        column_indices: list[int] | None = None,
        evaluator: ExcelFormulaEvaluator | None = None,
    ) -> list[list[object]]:
        rows: list[list[object]] = []
        if evaluator is None:
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True):
                if column_indices is None:
                    rows.append([None if cls._is_empty_cell(value) else value for value in row])
                    continue
                rows.append(
                    [
                        None if index >= len(row) or cls._is_empty_cell(row[index]) else row[index]
                        for index in column_indices
                    ]
                )
            return rows

        selected_indices = column_indices or list(range(max_column))
        for row_number in range(1, max_row + 1):
            row_values: list[object] = []
            for column_index in selected_indices:
                if column_index >= max_column:
                    row_values.append(None)
                    continue
                value = (
                    evaluator.cell_value(worksheet.title, row_number, column_index + 1)
                    if evaluator is not None
                    else worksheet.cell(row=row_number, column=column_index + 1).value
                )
                row_values.append(None if cls._is_empty_cell(value) else value)
            rows.append(row_values)
        return rows

    @classmethod
    def _worksheet_has_formulas(
        cls,
        worksheet: Any,
        *,
        max_row: int,
        max_column: int,
        column_indices: list[int] | None = None,
    ) -> bool:
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True):
            values = row if column_indices is None else [row[index] if index < len(row) else None for index in column_indices]
            if any(isinstance(value, str) and value.startswith("=") for value in values):
                return True
        return False

    @classmethod
    def _active_column_indices(cls, rows: list[list[object]]) -> list[int]:
        if not rows:
            return []
        max_columns = max(len(row) for row in rows)
        return [
            column_index
            for column_index in range(max_columns)
            if any(column_index < len(row) and not cls._is_empty_cell(row[column_index]) for row in rows)
        ]

    @staticmethod
    def _is_empty_cell(value: object) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        return False

    @staticmethod
    def _sample_rows(rows: list[list[str]], columns: list[str]) -> list[list[str]]:
        if len(rows) <= _MAX_SHEET_ROWS:
            return rows
        # 均匀采样：保留首部、中部、尾部，覆盖全表特征
        half = _MAX_SHEET_ROWS // 2
        return rows[:half] + rows[-half:]

    @staticmethod
    def _cell_text(value: object) -> str:
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        text = "" if value is None else str(value).strip()
        return normalize_whitespace(text)
