from __future__ import annotations

from collections.abc import Iterable, Iterator
from datetime import date
from typing import Any, cast

from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_to_tuple,
    range_boundaries,
)


class ExcelFormulaEvaluator:
    """Small, bounded evaluator for common report formulas.

    This is not a full Excel engine. It fills the practical gap where generated
    workbooks contain formulas but no cached values, so data_only reads return
    blanks for cells that the agent still needs to analyze.
    """

    def __init__(self, *, value_workbook: Any, formula_workbook: Any) -> None:
        self._value_workbook = value_workbook
        self._formula_workbook = formula_workbook
        self._cache: dict[tuple[str, int, int], object] = {}
        self._evaluating: set[tuple[str, int, int]] = set()
        self._value_rows_by_sheet: dict[str, _WorksheetRows] = {}
        self._formula_rows_by_sheet: dict[str, _WorksheetRows] = {}

    def cell_value(self, sheet_name: str, row: int, column: int) -> object:
        key = (sheet_name, row, column)
        if key in self._cache:
            return self._cache[key]
        if key in self._evaluating:
            return None

        self._evaluating.add(key)
        try:
            cached = self._value_rows(sheet_name).cell(row, column)
            if not _is_empty(cached):
                self._cache[key] = cached
                return cached

            raw = self._formula_rows(sheet_name).cell(row, column)
            if isinstance(raw, str) and raw.startswith("="):
                value = self._evaluate_formula(raw, sheet_name)
            else:
                value = raw
            self._cache[key] = value
            return value
        finally:
            self._evaluating.discard(key)

    def _evaluate_formula(self, formula: str, current_sheet: str) -> object:
        try:
            expression = self._to_python_expression(formula)
            return eval(  # noqa: S307 - expression is generated from openpyxl tokens with a closed environment.
                expression,
                {"__builtins__": {}},
                self._evaluation_env(current_sheet),
            )
        except Exception:
            return None

    @staticmethod
    def _to_python_expression(formula: str) -> str:
        pieces: list[str] = []
        for token in Tokenizer(formula).items:
            if token.type == "OPERAND" and token.subtype == "RANGE":
                pieces.append(f"REF({token.value!r})")
            elif token.type == "OPERAND" and token.subtype == "TEXT":
                pieces.append(token.value)
            elif token.type == "OPERAND" and token.subtype == "NUMBER":
                pieces.append(token.value)
            elif token.type == "OPERAND" and token.subtype == "LOGICAL":
                pieces.append("True" if token.value.upper() == "TRUE" else "False")
            elif token.type == "FUNC":
                pieces.append(token.value)
            elif token.type == "OPERATOR-INFIX":
                pieces.append(_python_operator(token.value))
            elif token.type in {"PAREN", "SEP"}:
                pieces.append(token.value)
            elif token.type == "OPERATOR-PREFIX":
                pieces.append(token.value)
        return "".join(pieces)

    def _evaluation_env(self, current_sheet: str) -> dict[str, object]:
        def _ref(reference: str) -> object:
            return self._resolve_reference(reference, current_sheet)

        return {
            "REF": _ref,
            "SUM": _sum_values,
            "SUMIF": _sumif,
            "VLOOKUP": _vlookup,
            "IF": lambda condition, true_value, false_value=None: true_value if bool(condition) else false_value,
            "IFERROR": lambda value, fallback=None: fallback if _is_empty(value) else value,
            "AND": lambda *values: all(bool(value) for value in values),
            "OR": lambda *values: any(bool(value) for value in values),
            "DATE": lambda year, month, day: date(int(year), int(month), int(day)),
            "TEXT": _text,
            "ABS": lambda value: abs(_number(value)),
            "ROUND": lambda value, ndigits=0: round(_number(value), int(_number(ndigits))),
            "MAX": lambda *values: max((_number(value) for value in _flatten(values)), default=0.0),
            "MIN": lambda *values: min((_number(value) for value in _flatten(values)), default=0.0),
        }

    def _resolve_reference(self, reference: str, current_sheet: str) -> object:
        sheet_name, address = _split_reference(reference, current_sheet)
        if ":" in address:
            min_col, min_row, max_col, max_row = _range_bounds(address, self._formula_workbook[sheet_name])
            return [
                [
                    self.cell_value(sheet_name, row=row, column=column)
                    for column in range(min_col, max_col + 1)
                ]
                for row in range(min_row, max_row + 1)
            ]
        row, column = coordinate_to_tuple(address)
        return self.cell_value(sheet_name, row=row, column=column)

    def _value_rows(self, sheet_name: str) -> _WorksheetRows:
        return self._rows_for_sheet(
            workbook=self._value_workbook,
            cache=self._value_rows_by_sheet,
            sheet_name=sheet_name,
        )

    def _formula_rows(self, sheet_name: str) -> _WorksheetRows:
        return self._rows_for_sheet(
            workbook=self._formula_workbook,
            cache=self._formula_rows_by_sheet,
            sheet_name=sheet_name,
        )

    @staticmethod
    def _rows_for_sheet(
        *,
        workbook: Any,
        cache: dict[str, _WorksheetRows],
        sheet_name: str,
    ) -> _WorksheetRows:
        rows = cache.get(sheet_name)
        if rows is None:
            rows = _WorksheetRows(workbook[sheet_name])
            cache[sheet_name] = rows
        return rows


class _WorksheetRows:
    def __init__(self, worksheet: Any) -> None:
        self._worksheet = worksheet
        self._rows: list[tuple[object, ...]] = []
        self._iterator: Iterator[tuple[object, ...]] | None = None
        self._exhausted = False

    @property
    def max_row(self) -> int:
        return int(getattr(self._worksheet, "max_row", 0) or 0)

    @property
    def max_column(self) -> int:
        return int(getattr(self._worksheet, "max_column", 0) or 0)

    def cell(self, row: int, column: int) -> object:
        if row <= 0 or column <= 0:
            return None
        self._ensure_loaded(row)
        if row > len(self._rows):
            return None
        values = self._rows[row - 1]
        if column > len(values):
            return None
        return values[column - 1]

    def _ensure_loaded(self, row: int) -> None:
        if row <= len(self._rows) or self._exhausted:
            return
        iterator = self._row_iterator()
        while len(self._rows) < row:
            try:
                self._rows.append(tuple(next(iterator)))
            except StopIteration:
                self._exhausted = True
                return

    def _row_iterator(self) -> Iterator[tuple[object, ...]]:
        if self._iterator is None:
            self._iterator = iter(
                self._worksheet.iter_rows(
                    min_row=1,
                    max_row=self.max_row or None,
                    max_col=self.max_column or None,
                    values_only=True,
                )
            )
        return self._iterator


def _split_reference(reference: str, current_sheet: str) -> tuple[str, str]:
    ref = reference.replace("$", "")
    if "!" not in ref:
        return current_sheet, ref
    sheet_name, address = ref.rsplit("!", 1)
    sheet_name = sheet_name.strip()
    if sheet_name.startswith("'") and sheet_name.endswith("'"):
        sheet_name = sheet_name[1:-1].replace("''", "'")
    return sheet_name, address


def _range_bounds(address: str, worksheet: Any) -> tuple[int, int, int, int]:
    if _is_full_column_range(address):
        left, right = address.split(":", 1)
        return (
            column_index_from_string(left),
            1,
            column_index_from_string(right),
            int(worksheet.max_row or 1),
        )
    return cast(tuple[int, int, int, int], range_boundaries(address))


def _is_full_column_range(address: str) -> bool:
    parts = address.split(":")
    return len(parts) == 2 and all(part.isalpha() for part in parts)


def _python_operator(operator: str) -> str:
    mapping = {
        "=": "==",
        "<>": "!=",
        "^": "**",
        "&": "+",
    }
    return mapping.get(operator, operator)


def _sum_values(*values: object) -> float:
    return sum(_number(value) for value in _flatten(values))


def _sumif(range_values: object, criteria: object, sum_range: object | None = None) -> float:
    criteria_values = list(_flatten([range_values]))
    target_values = list(_flatten([sum_range if sum_range is not None else range_values]))
    total = 0.0
    for index, value in enumerate(criteria_values):
        if index >= len(target_values):
            break
        if _matches_criteria(value, criteria):
            total += _number(target_values[index])
    return total


def _vlookup(lookup_value: object, table: object, col_index: object, range_lookup: object = False) -> object:
    del range_lookup
    rows = table if isinstance(table, list) else []
    target_index = max(0, int(_number(col_index)) - 1)
    for row in rows:
        if not isinstance(row, list) or not row:
            continue
        if _normalize_compare(row[0]) == _normalize_compare(lookup_value):
            if target_index < len(row):
                return row[target_index]
            return None
    return None


def _matches_criteria(value: object, criteria: object) -> bool:
    if isinstance(criteria, str):
        stripped = criteria.strip()
        for operator in (">=", "<=", "<>", ">", "<", "="):
            if stripped.startswith(operator):
                target = stripped[len(operator):].strip()
                return _compare(value, target, operator)
    return _normalize_compare(value) == _normalize_compare(criteria)


def _compare(left: object, right: object, operator: str) -> bool:
    left_num = _try_number(left)
    right_num = _try_number(right)
    left_value: float | str = left_num if left_num is not None and right_num is not None else _normalize_compare(left)
    right_value: float | str = (
        right_num if left_num is not None and right_num is not None else _normalize_compare(right)
    )
    if operator in {"=", "=="}:
        return left_value == right_value
    if operator == "<>":
        return left_value != right_value
    if operator == ">":
        return bool(left_value > right_value)  # type: ignore[operator]
    if operator == "<":
        return bool(left_value < right_value)  # type: ignore[operator]
    if operator == ">=":
        return bool(left_value >= right_value)  # type: ignore[operator]
    if operator == "<=":
        return bool(left_value <= right_value)  # type: ignore[operator]
    return False


def _text(value: object, fmt: object) -> str:
    fmt_text = str(fmt).lower()
    if isinstance(value, date) and "yyyy" in fmt_text and "mm" in fmt_text and "dd" in fmt_text:
        return value.strftime("%Y-%m-%d")
    return str(value)


def _flatten(values: Iterable[object]) -> Iterable[object]:
    for value in values:
        if isinstance(value, list):
            yield from _flatten(value)
        else:
            yield value


def _number(value: object) -> float:
    parsed = _try_number(value)
    return 0.0 if parsed is None else parsed


def _try_number(value: object) -> float | None:
    if _is_empty(value):
        return 0.0
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, int | float):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("，", "").strip())
    except ValueError:
        return None


def _normalize_compare(value: object) -> str:
    if _is_empty(value):
        return ""
    return str(value).strip()


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


__all__ = ["ExcelFormulaEvaluator"]
