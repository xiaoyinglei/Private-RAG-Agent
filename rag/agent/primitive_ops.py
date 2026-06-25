"""Primitive operations for agent workspace interactions."""

from __future__ import annotations

import base64
import csv
import mimetypes
import os
from pathlib import Path
from typing import Any, Literal

from pydantic import BaseModel, Field, model_validator

from rag.agent.memory.store import MEMORY_DIR_NAME
from rag.agent.runner.python_runner import (
    PythonRunner,
    SeatbeltPythonRunner,
)
from rag.agent.workspace import WorkspaceRuntime

# ---------------------------------------------------------------------------
# I/O models
# ---------------------------------------------------------------------------

MAX_LIST_FILES = 200
MAX_GENERATED_FILES = 200
FILE_SAMPLE_BYTES = 4096
FileKind = Literal["directory", "text", "binary", "unknown"]
CellValue = str | int | float | bool | None
MAX_PROBE_ROWS = 200
MAX_PROBE_COLUMNS = 200
MAX_PROBE_TABLES = 20


class ListFilesInput(BaseModel):
    path: str = ""
    pattern: str | None = None
    limit: int = Field(default=MAX_LIST_FILES, ge=1, le=MAX_LIST_FILES)


class FileInfo(BaseModel):
    name: str
    path: str
    size: int
    is_dir: bool
    modified_at: float
    mime_type: str | None = None
    file_kind: FileKind = "unknown"
    is_binary: bool = False
    readable_as_text: bool = False
    capabilities: list[str] = Field(default_factory=list)


class ListFilesOutput(BaseModel):
    files: list[FileInfo]
    truncated: bool = False


MAX_READ_BYTES = 10_000_000  # 10MB hard ceiling


class ReadFileInput(BaseModel):
    path: str
    encoding: str = "utf-8"
    max_bytes: int = Field(default=1_000_000, ge=1, le=MAX_READ_BYTES)
    offset: int = Field(default=0, ge=0, description="Byte offset to start reading from")
    limit: int | None = Field(
        default=None,
        ge=1,
        description="Maximum number of bytes to read (overrides max_bytes if set)",
    )


class ReadFileOutput(BaseModel):
    path: str
    content: str
    truncated: bool
    size_bytes: int
    is_binary: bool = False
    encoding: str = "utf-8"


class WriteFileInput(BaseModel):
    path: str
    content: str
    encoding: str = "utf-8"
    overwrite: bool = False


class WriteFileOutput(BaseModel):
    path: str
    size_bytes: int


MAX_PYTHON_TIMEOUT = 120.0  # seconds, hard ceiling


class RunPythonInput(BaseModel):
    script_path: str = Field(
        default="",
        description="Path to a Python script to run. Leave empty if using code parameter.",
    )
    code: str = Field(
        default="",
        max_length=50000,
        description="Python code to execute directly. Leave empty if using script_path. At least one must be provided.",
    )
    args: list[str] = Field(default_factory=list)
    timeout_seconds: float = Field(default=30.0, gt=0, le=MAX_PYTHON_TIMEOUT)

    @model_validator(mode="after")
    def _require_script_or_code(self) -> "RunPythonInput":
        if not self.script_path and not self.code:
            raise ValueError("Either script_path or code must be provided")
        return self


class RunPythonInlineInput(BaseModel):
    code: str = Field(
        min_length=1,
        max_length=50000,
        description="Python code to execute directly. Use for data analysis, file reading, computation.",
    )
    timeout_seconds: float = Field(default=30.0, gt=0, le=MAX_PYTHON_TIMEOUT)


class ImagePreview(BaseModel):
    """Preview of a generated image file."""
    path: str
    base64_data: str
    mime_type: str
    width: int | None = None
    height: int | None = None


class RunPythonOutput(BaseModel):
    ok: bool
    exit_code: int
    stdout: str
    stderr: str
    stdout_truncated: bool
    stderr_truncated: bool
    duration_ms: float
    generated_files: list[str]
    image_previews: list[ImagePreview] = Field(default_factory=list)


class StructuredProbeInput(BaseModel):
    path: str
    encoding: str = "utf-8"
    max_rows: int = Field(default=20, ge=1, le=MAX_PROBE_ROWS)
    max_columns: int = Field(default=50, ge=1, le=MAX_PROBE_COLUMNS)
    max_tables: int = Field(default=5, ge=1, le=MAX_PROBE_TABLES)


class CandidateHeaderRow(BaseModel):
    row_index: int = Field(ge=1)
    confidence: float = Field(ge=0.0, le=1.0)
    reason: str


class StructuredTableProbe(BaseModel):
    table_index: int = Field(ge=0)
    name: str | None = None
    used_range: str | None = None
    row_count: int = Field(ge=0)
    column_count: int = Field(ge=0)
    sample_rows: list[list[CellValue]] = Field(default_factory=list)
    candidate_header_rows: list[CandidateHeaderRow] = Field(default_factory=list)
    data_start_row: int | None = Field(default=None, ge=1)
    merged_cells: bool = False
    formula_columns: list[str] = Field(default_factory=list)


class StructuredProbeOutput(BaseModel):
    path: str
    file_kind: FileKind = "unknown"
    mime_type: str | None = None
    tables: list[StructuredTableProbe] = Field(default_factory=list)
    truncated: bool = False
    errors: list[str] = Field(default_factory=list)


# ---------------------------------------------------------------------------
# PrimitiveOps
# ---------------------------------------------------------------------------

class PrimitiveOps:
    _WRITABLE_DIRS = ("scratch", "artifacts", "reports", "logs")

    def __init__(
        self,
        workspace: WorkspaceRuntime,
        python_runner: PythonRunner | None = None,
    ) -> None:
        self._workspace = workspace
        self._python_runner = python_runner or SeatbeltPythonRunner()

    # -- list_files --------------------------------------------------------

    def list_files(self, payload: ListFilesInput) -> ListFilesOutput:
        base = (
            self._workspace.resolve_path(payload.path)
            if payload.path
            else self._workspace.root
        )
        self._workspace.ensure_within_workspace(base)
        if base.resolve() != self._workspace.root.resolve():
            _ensure_not_agent_memory(self._workspace, base)

        if not base.is_dir():
            return ListFilesOutput(files=[])

        entries: list[FileInfo] = []
        truncated = False
        for entry in sorted(base.iterdir()):
            if entry.name == MEMORY_DIR_NAME:
                continue
            if payload.pattern and not entry.match(payload.pattern):
                continue
            if len(entries) >= payload.limit:
                truncated = True
                break
            stat = entry.stat()
            rel = self._workspace.relative_to_root(entry)
            entries.append(
                _file_info(
                    entry,
                    rel=rel,
                    size=stat.st_size,
                    modified_at=stat.st_mtime,
                )
            )
        return ListFilesOutput(files=entries, truncated=truncated)

    # -- read_file ---------------------------------------------------------

    def read_file(self, payload: ReadFileInput) -> ReadFileOutput:
        target = self._workspace.resolve_path(payload.path)
        self._workspace.ensure_within_workspace(target)
        _ensure_not_agent_memory(self._workspace, target)

        if not target.is_file():
            raise FileNotFoundError(f"File not found: {payload.path}")

        size = target.stat().st_size
        # Effective byte limit: limit overrides max_bytes if set
        effective_max = (
            min(payload.limit, payload.max_bytes)
            if payload.limit is not None
            else payload.max_bytes
        )
        # Bounded read with offset support
        with target.open("rb") as f:
            if payload.offset > 0:
                f.seek(payload.offset)
            raw = f.read(effective_max + 1)
        truncated = len(raw) > effective_max
        if truncated:
            raw = raw[: effective_max]

        if _is_binary_content(
            raw,
            encoding=payload.encoding,
            mime_type=mimetypes.guess_type(target.name)[0],
        ):
            return ReadFileOutput(
                path=payload.path,
                content="",
                truncated=truncated,
                size_bytes=size,
                is_binary=True,
                encoding=payload.encoding,
            )

        content = raw.decode(payload.encoding)

        return ReadFileOutput(
            path=payload.path,
            content=content,
            truncated=truncated,
            size_bytes=size,
            encoding=payload.encoding,
        )

    # -- write_file --------------------------------------------------------

    def write_file(self, payload: WriteFileInput) -> WriteFileOutput:
        target = self._workspace.resolve_path(payload.path)
        self._workspace.ensure_within_workspace(target)

        rel = self._workspace.relative_to_root(target)
        top_dir = rel.parts[0] if rel.parts else ""
        if top_dir not in self._WRITABLE_DIRS:
            raise PermissionError(
                f"Cannot write to {top_dir or 'workspace root'}/: only "
                f"{', '.join(d + '/' for d in self._WRITABLE_DIRS)} are allowed"
            )

        if target.exists() and not payload.overwrite:
            raise FileExistsError(
                f"File already exists and overwrite=False: {payload.path}"
            )

        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(payload.content, encoding=payload.encoding)

        return WriteFileOutput(
            path=payload.path,
            size_bytes=len(payload.content.encode(payload.encoding)),
        )

    # -- run_python --------------------------------------------------------

    def run_python(self, payload: RunPythonInput) -> RunPythonOutput:
        # Code path: write injected preamble + user code to temp file
        if payload.code:
            return self._run_python_code(payload)

        # Script path: existing behaviour
        script_abs = self._workspace.resolve_path(payload.script_path)
        self._workspace.ensure_within_scratch(script_abs)

        if not script_abs.suffix == ".py":
            raise ValueError(f"Only .py files allowed: {payload.script_path}")

        if not script_abs.is_file():
            raise FileNotFoundError(f"Script not found: {payload.script_path}")

        return self._execute_python_file(script_abs, payload.args, payload.timeout_seconds)

    def _run_python_code(self, payload: RunPythonInput) -> RunPythonOutput:
        """Execute Python code with SDK preamble injected.

        Writes code to a temp .py file in scratch/, prepends the Tool SDK
        so agent-written Python can use tools.declare().
        """
        import tempfile

        # Resolve scratch dir (must be first — used by _matplotlib_preamble below)
        scratch = self._workspace.root / "scratch"
        scratch.mkdir(parents=True, exist_ok=True)

        # Load SDK preamble
        sdk_source = _load_sdk_preamble()

        # Build the full script: matplotlib preamble + SDK + user code
        matplot_preamble = _matplotlib_preamble(scratch)
        full_code = f"{matplot_preamble}\n\n{sdk_source}\n\n{payload.code}"

        # Write to temp file in scratch/
        with tempfile.NamedTemporaryFile(
            mode="w",
            suffix=".py",
            dir=str(scratch),
            delete=False,
            encoding="utf-8",
        ) as tf:
            tf.write(full_code)
            temp_path = Path(tf.name)

        # Set AGENT_SCRATCH_DIR so the SDK knows where to write tool_calls.jsonl
        prev_scratch = os.environ.get("AGENT_SCRATCH_DIR")
        try:
            os.environ["AGENT_SCRATCH_DIR"] = str(scratch)
            return self._execute_python_file(
                temp_path, payload.args, payload.timeout_seconds,
            )
        finally:
            if prev_scratch is not None:
                os.environ["AGENT_SCRATCH_DIR"] = prev_scratch
            else:
                os.environ.pop("AGENT_SCRATCH_DIR", None)
            # Clean up temp file
            try:
                temp_path.unlink(missing_ok=True)
            except OSError:
                pass

    def _execute_python_file(
        self,
        script_path: Path,
        args: list[str],
        timeout: float,
    ) -> RunPythonOutput:
        """Execute a Python file and capture results."""
        before = _snapshot_files(self._workspace.root)

        result = self._python_runner.run(
            script_path,
            args=args,
            cwd=self._workspace.root,
            timeout=timeout,
        )

        after = _snapshot_files(self._workspace.root)
        generated = sorted(after - before)
        generated_files = [
            str(self._workspace.relative_to_root(Path(f)))
            for f in generated[:MAX_GENERATED_FILES]
        ]

        image_previews = _collect_image_previews(
            [Path(f) for f in generated],
            workspace=self._workspace,
        )

        return RunPythonOutput(
            ok=result.exit_code == 0,
            exit_code=result.exit_code,
            stdout=result.stdout,
            stderr=result.stderr,
            stdout_truncated=len(result.stdout) >= 100_000,
            stderr_truncated=len(result.stderr) >= 50_000,
            duration_ms=result.duration_ms,
            generated_files=generated_files,
            image_previews=image_previews,
        )

    def run_python_inline(self, payload: RunPythonInlineInput) -> RunPythonOutput:
        """Execute Python code directly without writing to a file first.

        Automatically patches matplotlib.pyplot.show() to save figures
        to scratch/ so chart output is captured.
        """
        import tempfile
        scratch = self._workspace.root / "scratch"
        scratch.mkdir(exist_ok=True)

        # Prepend matplotlib auto-capture preamble
        preamble = _matplotlib_preamble(scratch)
        full_code = preamble + "\n" + payload.code

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".py", dir=scratch, delete=False,
        ) as f:
            f.write(full_code)
            script_path = Path(f.name)
        try:
            return self.run_python(RunPythonInput(
                script_path=str(self._workspace.relative_to_root(script_path)),
                timeout_seconds=payload.timeout_seconds,
            ))
        finally:
            script_path.unlink(missing_ok=True)

    # -- structured_probe --------------------------------------------------

    def structured_probe(self, payload: StructuredProbeInput) -> StructuredProbeOutput:
        target = self._workspace.resolve_path(payload.path)
        self._workspace.ensure_within_workspace(target)
        _ensure_not_agent_memory(self._workspace, target)

        if not target.is_file():
            raise FileNotFoundError(f"File not found: {payload.path}")

        stat = target.stat()
        rel = self._workspace.relative_to_root(target)
        info = _file_info(target, rel=rel, size=stat.st_size, modified_at=stat.st_mtime)
        errors: list[str] = []
        tables: list[StructuredTableProbe] = []
        truncated = False

        if _is_excel_file(target, info.mime_type):
            try:
                tables, truncated = _probe_excel_file(
                    target,
                    max_rows=payload.max_rows,
                    max_columns=payload.max_columns,
                    max_tables=payload.max_tables,
                )
            except Exception as exc:
                errors.append(f"excel_probe_failed: {exc}")
        elif _is_delimited_text_file(target, info.mime_type, readable_as_text=info.readable_as_text):
            try:
                table = _probe_delimited_file(
                    target,
                    encoding=payload.encoding,
                    max_rows=payload.max_rows,
                    max_columns=payload.max_columns,
                )
                tables = [table]
            except Exception as exc:
                errors.append(f"delimited_probe_failed: {exc}")
        else:
            errors.append("unsupported_structured_probe_type")

        return StructuredProbeOutput(
            path=payload.path,
            file_kind=info.file_kind,
            mime_type=info.mime_type,
            tables=tables,
            truncated=truncated,
            errors=errors,
        )

    # -- runners registry --------------------------------------------------

    def runners(self) -> dict[str, Any]:
        return {
            "list_files": self.list_files,
            "read_file": self.read_file,
            "structured_probe": self.structured_probe,
            "write_file": self.write_file,
            "run_python": self.run_python,
            "run_python_inline": self.run_python_inline,
            "tool_repl": self.tool_repl,
        }

    def tool_repl(self, payload: Any) -> RunPythonOutput:
        """Execute tool_repl — batch tool calling via Python code.

        Adapts RunCommandInput.command → RunPythonInput.code.
        """
        from rag.agent.tools.generic_tools import RunCommandInput

        if isinstance(payload, RunCommandInput):
            code = payload.command
        elif isinstance(payload, dict):
            code = payload.get("command", "")
        else:
            code = str(getattr(payload, "command", ""))
        return self.run_python(RunPythonInput(code=code))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _ensure_not_agent_memory(workspace: WorkspaceRuntime, path: Path) -> None:
    rel = workspace.relative_to_root(path)
    if rel.parts and rel.parts[0] == MEMORY_DIR_NAME:
        raise PermissionError(
            "Cannot access agent memory through primitive file tools; "
            "use internal MemoryStore.resolve(ref)"
        )


def _snapshot_files(root: Path) -> set[str]:
    files: set[str] = set()
    for path in root.rglob("*"):
        if path.is_file():
            files.add(str(path))
    return files


def _file_info(path: Path, *, rel: Path, size: int, modified_at: float) -> FileInfo:
    if path.is_dir():
        return FileInfo(
            name=path.name,
            path=str(rel),
            size=size,
            is_dir=True,
            modified_at=modified_at,
            file_kind="directory",
            capabilities=["list_files"],
        )

    mime_type = mimetypes.guess_type(path.name)[0]
    sample = _read_file_sample(path)
    is_binary = _is_binary_content(sample, encoding="utf-8", mime_type=mime_type)
    readable_as_text = not is_binary
    return FileInfo(
        name=path.name,
        path=str(rel),
        size=size,
        is_dir=False,
        modified_at=modified_at,
        mime_type=mime_type,
        file_kind="binary" if is_binary else "text",
        is_binary=is_binary,
        readable_as_text=readable_as_text,
        capabilities=(
            ["read_file", "structured_probe", "run_python"]
            if readable_as_text
            else ["structured_probe", "run_python"]
        ),
    )


def _read_file_sample(path: Path) -> bytes:
    try:
        with path.open("rb") as f:
            return f.read(FILE_SAMPLE_BYTES)
    except OSError:
        return b""


def _is_binary_content(
    raw: bytes,
    *,
    encoding: str,
    mime_type: str | None,
) -> bool:
    if mime_type is not None and _mime_type_is_binary(mime_type):
        return True

    if not raw:
        return False

    try:
        decoded = raw.decode(encoding)
    except UnicodeDecodeError:
        return True

    if "\x00" in decoded:
        return True

    control_chars = sum(
        1
        for ch in decoded
        if ord(ch) < 32 and ch not in "\n\r\t\f\b"
    )
    return control_chars / max(len(decoded), 1) > 0.05


def _mime_type_is_binary(mime_type: str) -> bool:
    normalized = mime_type.lower()
    if normalized.startswith("text/"):
        return False
    if normalized in {
        "application/json",
        "application/x-json",
        "application/xml",
        "application/x-yaml",
        "application/yaml",
        "application/toml",
    }:
        return False
    if normalized.endswith("+json") or normalized.endswith("+xml"):
        return False
    return True


def _is_excel_file(path: Path, mime_type: str | None) -> bool:
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return True
    return mime_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    }


def _is_delimited_text_file(
    path: Path,
    mime_type: str | None,
    *,
    readable_as_text: bool,
) -> bool:
    if path.suffix.lower() in {".csv", ".tsv"}:
        return True
    return readable_as_text and mime_type in {"text/csv", "text/tab-separated-values"}


def _probe_delimited_file(
    path: Path,
    *,
    encoding: str,
    max_rows: int,
    max_columns: int,
) -> StructuredTableProbe:
    dialect = _detect_csv_dialect(path, encoding=encoding)
    sample_rows: list[list[CellValue]] = []
    row_count = 0
    column_count = 0
    with path.open("r", encoding=encoding, newline="") as f:
        reader = csv.reader(f, dialect=dialect)
        for row in reader:
            row_count += 1
            column_count = max(column_count, len(row))
            if len(sample_rows) < max_rows:
                sample_rows.append([_cell_value(value) for value in row[:max_columns]])

    candidates = _candidate_header_rows(sample_rows)
    data_start_row = candidates[0].row_index + 1 if candidates else None
    return StructuredTableProbe(
        table_index=0,
        name=path.name,
        used_range=_used_range(row_count=row_count, column_count=column_count),
        row_count=row_count,
        column_count=column_count,
        sample_rows=sample_rows,
        candidate_header_rows=candidates,
        data_start_row=data_start_row,
    )


def _detect_csv_dialect(path: Path, *, encoding: str) -> type[csv.Dialect] | csv.Dialect:
    if path.suffix.lower() == ".tsv":
        return csv.excel_tab
    with path.open("r", encoding=encoding, newline="") as f:
        sample = f.read(FILE_SAMPLE_BYTES)
    try:
        return csv.Sniffer().sniff(sample)
    except csv.Error:
        return csv.excel


def _probe_excel_file(
    path: Path,
    *,
    max_rows: int,
    max_columns: int,
    max_tables: int,
) -> tuple[list[StructuredTableProbe], bool]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to probe Excel workbooks") from exc

    # First pass: data_only for values
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tables: list[StructuredTableProbe] = []
    for table_index, sheet_name in enumerate(workbook.sheetnames[:max_tables]):
        sheet = workbook[sheet_name]
        row_count = int(sheet.max_row or 0)
        column_count = int(sheet.max_column or 0)
        sample_rows: list[list[CellValue]] = []
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(max_rows, row_count),
            min_col=1,
            max_col=min(max_columns, column_count),
            values_only=True,
        ):
            sample_rows.append([_cell_value(value) for value in row])
        candidates = _candidate_header_rows(sample_rows)
        tables.append(
            StructuredTableProbe(
                table_index=table_index,
                name=str(sheet_name),
                used_range=sheet.calculate_dimension(),
                row_count=row_count,
                column_count=column_count,
                sample_rows=sample_rows,
                candidate_header_rows=candidates,
                data_start_row=candidates[0].row_index + 1 if candidates else None,
            )
        )
    workbook.close()

    # Second pass: detect merged cells and formulas (requires non-read-only)
    try:
        wb2 = openpyxl.load_workbook(path, read_only=False, data_only=False)
        for table in tables:
            if table.name and table.name in wb2.sheetnames:
                sheet = wb2[table.name]
                # Merged cells
                table.merged_cells = len(sheet.merged_cells.ranges) > 0
                # Formula detection: check first 10 rows
                formula_cols: list[str] = []
                max_col = min(table.column_count, max_columns)
                for row in sheet.iter_rows(
                    min_row=1, max_row=min(10, table.row_count),
                    max_col=max_col, values_only=False,
                ):
                    for cell in row:
                        if (cell.value and isinstance(cell.value, str)
                                and cell.value.startswith("=")):
                            col_letter = cell.column_letter
                            if col_letter not in formula_cols:
                                formula_cols.append(col_letter)
                table.formula_columns = formula_cols
        wb2.close()
    except Exception:
        pass  # Best-effort; merged_cells/formula_columns stay default

    return tables, len(workbook.sheetnames) > max_tables


def _candidate_header_rows(rows: list[list[CellValue]]) -> list[CandidateHeaderRow]:
    candidates: list[CandidateHeaderRow] = []
    for index, row in enumerate(rows):
        non_empty = [_normalized_cell(cell) for cell in row if _normalized_cell(cell)]
        if len(non_empty) < 2:
            continue
        label_like = [value for value in non_empty if _looks_like_label(value)]
        label_ratio = len(label_like) / len(non_empty)
        unique_ratio = len({value.lower() for value in label_like}) / max(len(label_like), 1)
        following = rows[index + 1 : index + 4]
        following_density = _following_density(following, expected=len(non_empty))
        data_evidence = 1.0 if _following_rows_look_like_data(following) else 0.0
        confidence = min(
            1.0,
            (0.45 * label_ratio)
            + (0.20 * unique_ratio)
            + (0.20 * following_density)
            + (0.15 * data_evidence),
        )
        if confidence < 0.55:
            continue
        candidates.append(
            CandidateHeaderRow(
                row_index=index + 1,
                confidence=round(confidence, 3),
                reason="label-like row followed by similarly shaped data rows",
            )
        )
    return sorted(candidates, key=lambda item: item.confidence, reverse=True)[:3]


def _following_density(rows: list[list[CellValue]], *, expected: int) -> float:
    if not rows or expected <= 0:
        return 0.0
    densities: list[float] = []
    for row in rows:
        non_empty = sum(1 for cell in row if _normalized_cell(cell))
        densities.append(min(1.0, non_empty / expected))
    return sum(densities) / len(densities)


def _following_rows_look_like_data(rows: list[list[CellValue]]) -> bool:
    for row in rows:
        non_empty = [cell for cell in row if _normalized_cell(cell)]
        if not non_empty:
            continue
        if any(isinstance(cell, int | float) and not isinstance(cell, bool) for cell in non_empty):
            return True
        if any(not _looks_like_label(_normalized_cell(cell)) for cell in non_empty):
            return True
    return False


def _looks_like_label(value: str) -> bool:
    stripped = value.strip()
    if not stripped or len(stripped) > 80:
        return False
    try:
        float(stripped.replace(",", ""))
    except ValueError:
        return True
    return False


def _normalized_cell(value: CellValue) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int | float):
        return str(value)
    return str(value).strip()


def _cell_value(value: object) -> CellValue:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        return str(isoformat())
    return str(value)


def _used_range(*, row_count: int, column_count: int) -> str | None:
    if row_count <= 0 or column_count <= 0:
        return None
    return f"A1:{_column_letter(column_count)}{row_count}"


def _column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


# ---------------------------------------------------------------------------
# Image / chart helpers
# ---------------------------------------------------------------------------

_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".svg", ".webp", ".gif", ".bmp"}

_IMAGE_MIME_MAP = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".svg": "image/svg+xml",
    ".webp": "image/webp",
    ".gif": "image/gif",
    ".bmp": "image/bmp",
}

MAX_IMAGE_PREVIEW_BYTES = 2_000_000  # 2MB


def _matplotlib_preamble(scratch_dir: Path) -> str:
    """Code injected before user code to auto-capture matplotlib figures."""
    scratch_str = str(scratch_dir).replace("\\", "\\\\")
    return f'''# -- matplotlib auto-capture preamble --
import os
os.makedirs("{scratch_str}", exist_ok=True)
try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    _plt_show_orig = plt.show
    _plt_save_counter = [0]
    def _plt_show_auto(*args, **kwargs):
        fig = plt.gcf()
        if fig.get_axes():
            _plt_save_counter[0] += 1
            path = os.path.join("{scratch_str}", f"chart_{{_plt_save_counter[0]:03d}}.png")
            fig.savefig(path, dpi=150, bbox_inches="tight")
        _plt_show_orig(*args, **kwargs)
    plt.show = _plt_show_auto
except ImportError:
    pass
# -- end preamble --
'''


_SDK_PREAMBLE_CACHE: str | None = None


def _load_sdk_preamble() -> str:
    """Load the Tool SDK source as a string to prepend to user code."""
    global _SDK_PREAMBLE_CACHE
    if _SDK_PREAMBLE_CACHE is not None:
        return _SDK_PREAMBLE_CACHE
    sdk_path = Path(__file__).resolve().parent / "tool_sdk.py"
    if sdk_path.exists():
        _SDK_PREAMBLE_CACHE = sdk_path.read_text(encoding="utf-8")
        return _SDK_PREAMBLE_CACHE
    # Fallback: minimal inline SDK
    _SDK_PREAMBLE_CACHE = """
import json, os
class _ToolDeclarer:
    def __init__(self): self._count = 0
    def declare(self, name, **args):
        if self._count >= int(os.environ.get('AGENT_MAX_BATCH_SIZE','10')): return {'declared':False}
        batch=os.path.join(os.environ.get('AGENT_SCRATCH_DIR','.'),'tool_calls.jsonl')
        import pathlib; pathlib.Path(batch).parent.mkdir(parents=True,exist_ok=True)
        with open(batch,'a') as f: f.write(json.dumps({'tool_name':name,'arguments':args},default=str)+'\\n')
        self._count+=1; return {'declared':name,'seq':self._count}
    def list_available(self): return os.environ.get('AGENT_AVAILABLE_TOOLS','').split(',')
tools = _ToolDeclarer()
"""
    return _SDK_PREAMBLE_CACHE


def _collect_image_previews(
    file_paths: list[Path],
    *,
    workspace: WorkspaceRuntime,
) -> list[ImagePreview]:
    """Collect base64 previews for generated image files."""
    previews: list[ImagePreview] = []
    for path in file_paths:
        if not path.is_file():
            continue
        ext = path.suffix.lower()
        if ext not in _IMAGE_EXTENSIONS:
            continue
        size = path.stat().st_size
        if size > MAX_IMAGE_PREVIEW_BYTES:
            continue
        try:
            data = path.read_bytes()
            b64 = base64.b64encode(data).decode("ascii")
            mime = _IMAGE_MIME_MAP.get(ext, "image/png")
            rel_path = str(workspace.relative_to_root(path))
            # Try to get dimensions
            width, height = _image_dimensions(path)
            previews.append(ImagePreview(
                path=rel_path,
                base64_data=b64,
                mime_type=mime,
                width=width,
                height=height,
            ))
        except Exception:
            continue
    return previews


def _image_dimensions(path: Path) -> tuple[int | None, int | None]:
    """Get image dimensions using PIL if available."""
    try:
        from PIL import Image
        with Image.open(path) as img:
            return img.size
    except Exception:
        return None, None


__all__ = [
    "FileInfo",
    "ImagePreview",
    "ListFilesInput",
    "ListFilesOutput",
    "PrimitiveOps",
    "ReadFileInput",
    "ReadFileOutput",
    "RunPythonInput",
    "RunPythonOutput",
    "CandidateHeaderRow",
    "StructuredProbeInput",
    "StructuredProbeOutput",
    "StructuredTableProbe",
    "WriteFileInput",
    "WriteFileOutput",
]
