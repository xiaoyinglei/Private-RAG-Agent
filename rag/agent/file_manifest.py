"""File manifest for agent input file awareness.

Builds a typed manifest of all input files, including pandas-based
preview for structured files (CSV/XLSX), so the model has bounded
file context on the first turn.
"""

from __future__ import annotations

import csv
import hashlib
import logging
import mimetypes
from pathlib import Path
from typing import Any

from pydantic import BaseModel, Field

from rag.agent.primitive_ops import (
    CandidateHeaderRow,
    CellValue,
    StructuredProbeOutput,
    StructuredTableProbe,
)
from rag.agent.workspace import WorkspaceRuntime

logger = logging.getLogger(__name__)

# ── File kind detection ──

_STRUCTURED_EXTENSIONS: dict[str, str] = {
    ".csv": "csv",
    ".tsv": "tsv",
    ".xlsx": "xlsx",
    ".xls": "xls",
    ".xlsm": "xlsm",
    ".xltx": "xltx",
    ".xltm": "xltm",
    ".json": "json",
    ".jsonl": "jsonl",
    ".parquet": "parquet",
}


def _detect_file_kind(path: Path, mime_type: str | None) -> str:
    ext = path.suffix.lower()
    if ext in _STRUCTURED_EXTENSIONS:
        return _STRUCTURED_EXTENSIONS[ext]
    if mime_type:
        if "spreadsheet" in mime_type or "excel" in mime_type:
            return "xlsx"
        if "csv" in mime_type:
            return "csv"
        if "tab-separated" in mime_type:
            return "tsv"
        if "parquet" in mime_type:
            return "parquet"
    if ext in {".txt", ".md", ".log"}:
        return "text"
    if ext in {".pdf"}:
        return "pdf"
    if ext in {".docx", ".doc"}:
        return "docx"
    return "unknown"


def _is_structured_kind(kind: str) -> bool:
    return kind in {"csv", "tsv", "xlsx", "xls", "xlsm", "xltx", "xltm", "json", "jsonl", "parquet"}


def _can_probe(kind: str) -> bool:
    return kind in {"csv", "tsv", "xlsx", "xls", "xlsm", "xltx", "xltm"}


# ── Models ──


class ColumnPreview(BaseModel):
    """Column metadata from pandas preview."""

    name: str
    dtype: str


class SheetPreview(BaseModel):
    """Pandas-based preview for one sheet/table."""

    sheet_name: str
    total_rows: int
    total_columns: int
    columns: list[ColumnPreview]
    head: list[dict[str, Any]]  # first N rows as records
    dtypes: dict[str, str]  # column_name -> dtype string
    header_row: int | None = None  # 1-based, from probe
    merged_cells: bool = False
    formula_columns: list[str] = Field(default_factory=list)


class FileManifestEntry(BaseModel):
    """Manifest for a single input file."""

    path: str  # workspace-relative path
    filename: str  # original filename
    size_bytes: int
    mime_type: str | None
    file_kind: str  # csv, xlsx, text, pdf, ...
    hash: str  # sha256 prefix for reference
    structured: bool  # can be analyzed with code
    probeable: bool  # supports a built-in structured preview
    sheets: list[SheetPreview] = Field(default_factory=list)
    probe: StructuredProbeOutput | None = None
    error: str | None = None  # if probe/preview failed


class FileManifest(BaseModel):
    """Manifest of all input files."""

    files: list[FileManifestEntry]
    total_size_bytes: int
    has_structured_files: bool
    has_probeable_files: bool

    def to_context_block(self) -> str:
        """Render manifest as a context block for the model."""
        if not self.files:
            return ""

        lines = ["── Input Files ──"]
        for i, entry in enumerate(self.files, 1):
            kind_label = entry.file_kind
            struct_label = ", structured" if entry.structured else ""
            lines.append(f"[{i}] {entry.path} ({_fmt_size(entry.size_bytes)}, {kind_label}{struct_label})")

            if entry.error:
                lines.append(f"    Error: {entry.error}")
                continue

            for sheet in entry.sheets:
                lines.append(f"    Sheet '{sheet.sheet_name}': {sheet.total_rows} rows × {sheet.total_columns} columns")
                if sheet.merged_cells:
                    lines.append("    WARNING: merged cells detected")
                if sheet.formula_columns:
                    lines.append(f"    WARNING: formulas in columns: {', '.join(sheet.formula_columns)}")
                if sheet.columns:
                    col_strs = [f"{c.name}({c.dtype})" for c in sheet.columns]
                    lines.append(f"    Columns: {', '.join(col_strs)}")
                if sheet.head:
                    lines.append(f"    Head({len(sheet.head)}):")
                    for row in sheet.head[:3]:  # max 3 rows in context
                        lines.append(f"      {row}")
                    if len(sheet.head) > 3:
                        lines.append(f"      ... ({len(sheet.head) - 3} more)")

            if entry.probe and not entry.sheets:
                # Fallback: show probe summary even without pandas preview
                for table in entry.probe.tables:
                    lines.append(f"    Table '{table.name}': {table.row_count} rows × {table.column_count} columns")
                    if table.merged_cells:
                        lines.append("    WARNING: merged cells detected")
                    if table.formula_columns:
                        lines.append(f"    WARNING: formulas in columns: {', '.join(table.formula_columns)}")
                    if table.candidate_header_rows:
                        best = table.candidate_header_rows[0]
                        lines.append(f"    Header candidate: row {best.row_index} (confidence {best.confidence:.2f})")

        # Available packages
        lines.append("")
        lines.append("── Available Python Packages ──")
        lines.append(_available_packages_str())

        return "\n".join(lines)


# ── Builder ──

# Pandas preview limits
_PREVIEW_ROWS = 5
_PROBE_MAX_ROWS = 10
_PROBE_MAX_COLUMNS = 50
_FILE_SAMPLE_BYTES = 4096


def build_file_manifest(workspace: WorkspaceRuntime) -> FileManifest:
    """Build manifest for all files in workspace/input_files/.

    For structured files (csv/xlsx), runs a lightweight schema and pandas
    preview to give the model enough context on the first turn.
    """
    input_dir = workspace.input_files
    if not input_dir.is_dir():
        return FileManifest(
            files=[],
            total_size_bytes=0,
            has_structured_files=False,
            has_probeable_files=False,
        )

    entries: list[FileManifestEntry] = []
    total_size = 0

    for path in sorted(input_dir.iterdir()):
        if not path.is_file():
            continue
        try:
            entry = _build_entry(path, workspace)
            entries.append(entry)
            total_size += entry.size_bytes
        except Exception as exc:
            logger.warning("Failed to build manifest for %s: %s", path, exc)
            stat = path.stat()
            entries.append(
                FileManifestEntry(
                    path=str(workspace.relative_to_root(path)),
                    filename=path.name,
                    size_bytes=stat.st_size,
                    mime_type=mimetypes.guess_type(path.name)[0],
                    file_kind="unknown",
                    hash="",
                    structured=False,
                    probeable=False,
                    error=str(exc),
                )
            )
            total_size += stat.st_size

    has_structured = any(e.structured for e in entries)
    has_probeable = any(e.probeable for e in entries)
    return FileManifest(
        files=entries,
        total_size_bytes=total_size,
        has_structured_files=has_structured,
        has_probeable_files=has_probeable,
    )


def _build_entry(path: Path, workspace: WorkspaceRuntime) -> FileManifestEntry:
    stat = path.stat()
    mime_type = mimetypes.guess_type(path.name)[0]
    file_kind = _detect_file_kind(path, mime_type)
    structured = _is_structured_kind(file_kind)
    probeable = _can_probe(file_kind)
    rel_path = str(workspace.relative_to_root(path))
    file_hash = _file_hash(path)

    sheets: list[SheetPreview] = []
    probe: StructuredProbeOutput | None = None
    error: str | None = None

    if probeable:
        # Build the bounded structured preview.
        try:
            probe = _run_probe(path, mime_type)
        except Exception as exc:
            error = f"probe_failed: {exc}"

        # Run pandas preview for probeable types
        if probe and not error:
            try:
                sheets = _build_pandas_previews(path, file_kind, probe)
            except Exception as exc:
                # Pandas preview is best-effort; probe data is still useful
                logger.debug("Pandas preview failed for %s: %s", path, exc)

    return FileManifestEntry(
        path=rel_path,
        filename=path.name,
        size_bytes=stat.st_size,
        mime_type=mime_type,
        file_kind=file_kind,
        hash=file_hash,
        structured=structured,
        probeable=probeable,
        sheets=sheets,
        probe=probe,
        error=error,
    )


def _run_probe(path: Path, mime_type: str | None) -> StructuredProbeOutput:
    """Build one bounded structured preview without using a Tool runtime."""

    errors: list[str] = []
    tables: list[StructuredTableProbe] = []
    truncated = False

    # Determine readable_as_text from file sample
    try:
        sample = path.open("rb").read(4096)
        is_binary = b"\x00" in sample
    except Exception:
        is_binary = True
    readable_as_text = not is_binary

    if _is_excel_file(path, mime_type):
        tables, truncated = _probe_excel_file(
            path,
            max_rows=_PROBE_MAX_ROWS,
            max_columns=_PROBE_MAX_COLUMNS,
            max_tables=20,
        )
    elif _is_delimited_text_file(path, mime_type, readable_as_text=readable_as_text):
        tables = [
            _probe_delimited_file(
                path,
                encoding="utf-8",
                max_rows=_PROBE_MAX_ROWS,
                max_columns=_PROBE_MAX_COLUMNS,
            )
        ]
    else:
        errors.append("unsupported_probe_type")

    return StructuredProbeOutput(
        path=str(path),
        file_kind="binary" if is_binary else "text",
        mime_type=mime_type,
        tables=tables,
        truncated=truncated,
        errors=errors,
    )


def _is_excel_file(path: Path, mime_type: str | None) -> bool:
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return True
    return mime_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    }


def _is_delimited_text_file(
    path: Path,
    mime_type: str | None,
    *,
    readable_as_text: bool,
) -> bool:
    if path.suffix.lower() in {".csv", ".tsv"}:
        return True
    return readable_as_text and mime_type in {
        "text/csv",
        "text/tab-separated-values",
    }


def _probe_delimited_file(
    path: Path,
    *,
    encoding: str,
    max_rows: int,
    max_columns: int,
) -> StructuredTableProbe:
    dialect = _detect_csv_dialect(path, encoding=encoding)
    sample_rows: list[list[CellValue]] = []
    row_count = 0
    column_count = 0
    with path.open("r", encoding=encoding, newline="") as file:
        reader = csv.reader(file, dialect=dialect)
        for row in reader:
            row_count += 1
            column_count = max(column_count, len(row))
            if len(sample_rows) < max_rows:
                sample_rows.append([_cell_value(value) for value in row[:max_columns]])

    candidates = _candidate_header_rows(sample_rows)
    data_start_row = candidates[0].row_index + 1 if candidates else None
    return StructuredTableProbe(
        table_index=0,
        name=path.name,
        used_range=_used_range(
            row_count=row_count,
            column_count=column_count,
        ),
        row_count=row_count,
        column_count=column_count,
        sample_rows=sample_rows,
        candidate_header_rows=candidates,
        data_start_row=data_start_row,
    )


def _detect_csv_dialect(
    path: Path,
    *,
    encoding: str,
) -> type[csv.Dialect] | csv.Dialect:
    if path.suffix.lower() == ".tsv":
        return csv.excel_tab
    with path.open("r", encoding=encoding, newline="") as file:
        sample = file.read(_FILE_SAMPLE_BYTES)
    try:
        return csv.Sniffer().sniff(sample)
    except csv.Error:
        return csv.excel


def _probe_excel_file(
    path: Path,
    *,
    max_rows: int,
    max_columns: int,
    max_tables: int,
) -> tuple[list[StructuredTableProbe], bool]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to preview Excel workbooks") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tables: list[StructuredTableProbe] = []
    for table_index, sheet_name in enumerate(workbook.sheetnames[:max_tables]):
        sheet = workbook[sheet_name]
        row_count = int(sheet.max_row or 0)
        column_count = int(sheet.max_column or 0)
        sample_rows: list[list[CellValue]] = []
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(max_rows, row_count),
            min_col=1,
            max_col=min(max_columns, column_count),
            values_only=True,
        ):
            sample_rows.append([_cell_value(value) for value in row])
        candidates = _candidate_header_rows(sample_rows)
        tables.append(
            StructuredTableProbe(
                table_index=table_index,
                name=str(sheet_name),
                used_range=sheet.calculate_dimension(),
                row_count=row_count,
                column_count=column_count,
                sample_rows=sample_rows,
                candidate_header_rows=candidates,
                data_start_row=(candidates[0].row_index + 1 if candidates else None),
            )
        )
    truncated = len(workbook.sheetnames) > max_tables
    workbook.close()

    try:
        workbook_with_formulas = openpyxl.load_workbook(
            path,
            read_only=False,
            data_only=False,
        )
        for table in tables:
            if table.name and table.name in workbook_with_formulas.sheetnames:
                sheet = workbook_with_formulas[table.name]
                table.merged_cells = bool(sheet.merged_cells.ranges)
                formula_columns: list[str] = []
                max_column = min(table.column_count, max_columns)
                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=min(10, table.row_count),
                    max_col=max_column,
                    values_only=False,
                ):
                    for cell in row:
                        if (
                            isinstance(cell.value, str)
                            and cell.value.startswith("=")
                            and cell.column_letter not in formula_columns
                        ):
                            formula_columns.append(cell.column_letter)
                table.formula_columns = formula_columns
        workbook_with_formulas.close()
    except Exception:
        pass

    return tables, truncated


def _candidate_header_rows(
    rows: list[list[CellValue]],
) -> list[CandidateHeaderRow]:
    candidates: list[CandidateHeaderRow] = []
    for index, row in enumerate(rows):
        non_empty = [_normalized_cell(cell) for cell in row if _normalized_cell(cell)]
        if len(non_empty) < 2:
            continue
        label_like = [value for value in non_empty if _looks_like_label(value)]
        label_ratio = len(label_like) / len(non_empty)
        unique_ratio = len({value.lower() for value in label_like}) / max(
            len(label_like),
            1,
        )
        following = rows[index + 1 : index + 4]
        following_density = _following_density(
            following,
            expected=len(non_empty),
        )
        data_evidence = 1.0 if _following_rows_look_like_data(following) else 0.0
        confidence = min(
            1.0,
            (0.45 * label_ratio) + (0.20 * unique_ratio) + (0.20 * following_density) + (0.15 * data_evidence),
        )
        if confidence < 0.55:
            continue
        candidates.append(
            CandidateHeaderRow(
                row_index=index + 1,
                confidence=round(confidence, 3),
                reason=("label-like row followed by similarly shaped data rows"),
            )
        )
    return sorted(
        candidates,
        key=lambda item: item.confidence,
        reverse=True,
    )[:3]


def _following_density(
    rows: list[list[CellValue]],
    *,
    expected: int,
) -> float:
    if not rows or expected <= 0:
        return 0.0
    densities: list[float] = []
    for row in rows:
        non_empty = sum(1 for cell in row if _normalized_cell(cell))
        densities.append(min(1.0, non_empty / expected))
    return sum(densities) / len(densities)


def _following_rows_look_like_data(rows: list[list[CellValue]]) -> bool:
    for row in rows:
        non_empty = [cell for cell in row if _normalized_cell(cell)]
        if not non_empty:
            continue
        if any(isinstance(cell, int | float) and not isinstance(cell, bool) for cell in non_empty):
            return True
        if any(not _looks_like_label(_normalized_cell(cell)) for cell in non_empty):
            return True
    return False


def _looks_like_label(value: str) -> bool:
    stripped = value.strip()
    if not stripped or len(stripped) > 80:
        return False
    try:
        float(stripped.replace(",", ""))
    except ValueError:
        return True
    return False


def _normalized_cell(value: CellValue) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int | float):
        return str(value)
    return str(value).strip()


def _cell_value(value: object) -> CellValue:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        return str(isoformat())
    return str(value)


def _used_range(*, row_count: int, column_count: int) -> str | None:
    if row_count <= 0 or column_count <= 0:
        return None
    return f"A1:{_column_letter(column_count)}{row_count}"


def _column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _build_pandas_previews(
    path: Path,
    file_kind: str,
    probe: StructuredProbeOutput,
) -> list[SheetPreview]:
    """Build pandas-based preview for each sheet/table."""
    sheets: list[SheetPreview] = []

    if file_kind in {"xlsx", "xls", "xlsm", "xltx", "xltm"}:
        sheets = _pandas_preview_excel(path, probe)
    elif file_kind in {"csv", "tsv"}:
        sheet = _pandas_preview_csv(path, file_kind, probe)
        if sheet:
            sheets.append(sheet)

    return sheets


def _pandas_preview_excel(
    path: Path,
    probe: StructuredProbeOutput,
) -> list[SheetPreview]:
    """Pandas preview for Excel sheets.

    Uses probe-detected header row so pandas reads the correct column names.
    """
    try:
        import pandas as pd
    except ImportError:
        return []

    sheets: list[SheetPreview] = []
    for table in probe.tables:
        sheet_name = table.name or f"Sheet{table.table_index + 1}"
        try:
            # Use probe's header detection to pick the right row
            header_row = None
            if table.candidate_header_rows:
                header_row = table.candidate_header_rows[0].row_index

            # pandas header param is 0-based; probe row_index is 1-based
            pandas_header = (header_row - 1) if header_row else 0
            # Skip rows before the header, then read header + data rows
            skip_rows = list(range(pandas_header)) if pandas_header > 0 else None

            df = pd.read_excel(
                path,
                sheet_name=sheet_name,
                header=pandas_header,
                skiprows=skip_rows,
                nrows=_PREVIEW_ROWS,
            )
            # Clean up column names (NaN → Unnamed)
            df.columns = [str(c) if not str(c).startswith("Unnamed") else f"col_{i}" for i, c in enumerate(df.columns)]
            columns = [ColumnPreview(name=c, dtype=str(df[c].dtype)) for c in df.columns]
            head = df.head(_PREVIEW_ROWS).to_dict(orient="records")
            dtypes = {c: str(df[c].dtype) for c in df.columns}

            # Detect merged cells (from openpyxl)
            merged_cells = _detect_merged_cells(path, sheet_name)

            # Detect formulas
            formula_cols = _detect_formula_columns(path, sheet_name, table.column_count)

            sheets.append(
                SheetPreview(
                    sheet_name=sheet_name,
                    total_rows=table.row_count,
                    total_columns=table.column_count,
                    columns=columns,
                    head=head,
                    dtypes=dtypes,
                    header_row=header_row,
                    merged_cells=merged_cells,
                    formula_columns=formula_cols,
                )
            )
        except Exception as exc:
            logger.debug("Pandas preview failed for sheet %s: %s", sheet_name, exc)

    return sheets


def _pandas_preview_csv(
    path: Path,
    file_kind: str,
    probe: StructuredProbeOutput,
) -> SheetPreview | None:
    """Pandas preview for CSV/TSV.

    Uses probe-detected header row so pandas reads the correct column names.
    """
    try:
        import pandas as pd
    except ImportError:
        return None

    sep = "\t" if file_kind == "tsv" else ","
    table = probe.tables[0] if probe.tables else None
    try:
        # Use probe's header detection
        header_row = None
        if table and table.candidate_header_rows:
            header_row = table.candidate_header_rows[0].row_index

        pandas_header = (header_row - 1) if header_row else 0
        skip_rows = list(range(pandas_header)) if pandas_header > 0 else None

        df = pd.read_csv(path, sep=sep, header=pandas_header, skiprows=skip_rows, nrows=_PREVIEW_ROWS)
        df.columns = [str(c) if not str(c).startswith("Unnamed") else f"col_{i}" for i, c in enumerate(df.columns)]
        columns = [ColumnPreview(name=c, dtype=str(df[c].dtype)) for c in df.columns]
        head = df.head(_PREVIEW_ROWS).to_dict(orient="records")
        dtypes = {c: str(df[c].dtype) for c in df.columns}

        return SheetPreview(
            sheet_name=path.name,
            total_rows=table.row_count if table else len(df),
            total_columns=table.column_count if table else len(df.columns),
            columns=columns,
            head=head,
            dtypes=dtypes,
            header_row=header_row,
        )
    except Exception as exc:
        logger.debug("Pandas CSV preview failed: %s", exc)
        return None


def _detect_merged_cells(path: Path, sheet_name: str) -> bool:
    """Check for merged cells using openpyxl."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=False)
        sheet = wb[sheet_name]
        has_merged = len(sheet.merged_cells.ranges) > 0
        wb.close()
        return has_merged
    except Exception:
        return False


def _detect_formula_columns(
    path: Path,
    sheet_name: str,
    max_cols: int,
) -> list[str]:
    """Detect columns containing formulas using openpyxl."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        sheet = wb[sheet_name]
        formula_cols: list[str] = []
        # Check first few rows for formulas
        for row in sheet.iter_rows(min_row=1, max_row=min(10, sheet.max_row or 1), max_col=max_cols, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    col_letter = cell.column_letter
                    if col_letter not in formula_cols:
                        formula_cols.append(col_letter)
        wb.close()
        return formula_cols
    except Exception:
        return []


def _file_hash(path: Path) -> str:
    """SHA-256 hash prefix for file reference."""
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def _fmt_size(size: int) -> str:
    if size < 1024:
        return f"{size} B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size / (1024 * 1024):.1f} MB"


def _available_packages_str() -> str:
    """Return a string of available Python packages for the model."""
    packages = []
    for pkg in ["pandas", "openpyxl", "matplotlib", "pillow", "numpy"]:
        try:
            mod = __import__(pkg)
            version = getattr(mod, "__version__", "")
            packages.append(f"{pkg} {version}" if version else pkg)
        except ImportError:
            pass
    return ", ".join(packages) if packages else "standard library only"


__all__ = [
    "ColumnPreview",
    "FileManifest",
    "FileManifestEntry",
    "SheetPreview",
    "build_file_manifest",
]
